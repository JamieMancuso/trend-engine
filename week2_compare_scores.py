"""
Week 2 Score Comparison — hand scores vs LLM scores
----------------------------------------------------
Joins the operator's hand-scored eval set against an LLM scoring output CSV
(from week2_run_scoring.py, or a proxy run) and produces a side-by-side
comparison spreadsheet showing where the two diverge.

Usage:
    py -3.14 week2_compare_scores.py \\
        --hand eval_set_v1__Scoring.csv \\
        --llm results_2026-04-18_174500.csv \\
        --out eval_v0.1_comparison.xlsx

Schema note: the operator's CSV uses "commercialization"; the v0.1 prompt
renamed this dimension to "profit_mechanism". This script treats them as
the same dimension for comparison purposes (per charter 2026-04-18 decision).
Divergence on that axis is expected and is itself a tuning signal.

Papers #2 and #3 are flagged as "over operator's head" per the charter and
are excluded from the primary calibration aggregates (but kept in the
per-paper tables as translation-layer test cases).
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# ---- CONFIG -----------------------------------------------------------------

# Papers the operator flagged as "over my head" (per charter 2026-04-18).
# Excluded from mean-divergence calculations but shown in the per-paper table.
OVER_HEAD_IDS = {2, 3}

# Stress-test papers (per charter 2026-04-18).
STRESS_TEST_IDS = {16, 17}

# Dimension pairs: (operator column, LLM column, label)
DIMS = [
    ("maturation",           "llm_maturation",           "maturation"),
    ("commercialization",    "llm_profit_mechanism",     "profit_mech"),
    ("retail_accessibility", "llm_retail_accessibility", "retail_access"),
    ("specificity",          "llm_specificity",          "specificity"),
    ("final",                "llm_final",                "final"),
]


# ---- IO ---------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--hand", default="eval_set_v1__Scoring.csv",
                   help="Operator's hand-scored eval CSV.")
    p.add_argument("--llm", required=True, help="LLM scoring output CSV.")
    p.add_argument("--out", default="eval_v0.1_comparison.xlsx",
                   help="Output xlsx path.")
    return p.parse_args()


def load_csv_by_id(path: str) -> dict[int, dict]:
    """Load a CSV keyed by integer id. Raises if rows lack an id."""
    out = {}
    with open(path, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            pid = int(row["id"])
            out[pid] = row
    return out


def to_float(x, default=None):
    """Best-effort float conversion; return default on blank/junk."""
    if x is None or x == "":
        return default
    try:
        return float(x)
    except (ValueError, TypeError):
        return default


# ---- ANALYTICS --------------------------------------------------------------

def compute_divergences(hand: dict, llm: dict) -> list[dict]:
    """Per-paper, per-dimension LLM minus hand. Returns list of row dicts
    ordered by paper id, ready to feed into the xlsx writer."""
    rows = []
    for pid in sorted(hand):
        if pid not in llm:
            continue
        hrow = hand[pid]
        lrow = llm[pid]

        row = {
            "id": pid,
            "domain": hrow.get("domain", ""),
            "title": hrow.get("title", ""),
            "hand_rationale": hrow.get("rationale", "").strip(),
            "llm_rationale": lrow.get("llm_rationale", "").strip(),
            "llm_flag": lrow.get("llm_flag", ""),
            "llm_time_to_thesis": lrow.get("llm_time_to_thesis", ""),
            "llm_translation": lrow.get("llm_translation", ""),
            "llm_public_vehicles": lrow.get("llm_public_vehicles", "[]"),
            "notes": [],
        }
        if pid in OVER_HEAD_IDS:
            row["notes"].append("over-operator's-head")
        if pid in STRESS_TEST_IDS:
            row["notes"].append("stress-test")

        for hand_col, llm_col, label in DIMS:
            h = to_float(hrow.get(hand_col))
            l = to_float(lrow.get(llm_col))
            row[f"hand_{label}"] = h
            row[f"llm_{label}"] = l
            row[f"delta_{label}"] = (l - h) if (h is not None and l is not None) else None

        row["notes"] = ", ".join(row["notes"])
        rows.append(row)
    return rows


def aggregate_stats(rows: list[dict], exclude_ids: set[int]) -> dict:
    """Mean absolute delta per dimension, ignoring excluded papers."""
    out = {}
    included = [r for r in rows if r["id"] not in exclude_ids]
    for _, _, label in DIMS:
        deltas = [r[f"delta_{label}"] for r in included if r[f"delta_{label}"] is not None]
        if not deltas:
            out[label] = {"mean_delta": None, "mean_abs_delta": None, "n": 0}
            continue
        mean_delta = sum(deltas) / len(deltas)
        mean_abs_delta = sum(abs(d) for d in deltas) / len(deltas)
        out[label] = {
            "mean_delta": mean_delta,
            "mean_abs_delta": mean_abs_delta,
            "n": len(deltas),
            "max_delta": max(deltas),
            "min_delta": min(deltas),
        }
    return out


def top_divergences(rows: list[dict], n: int = 5) -> list[tuple[int, float, str]]:
    """Papers with largest total absolute divergence across all 4 sub-dims + final."""
    scored = []
    for r in rows:
        total = 0.0
        missing = False
        for _, _, label in DIMS:
            d = r[f"delta_{label}"]
            if d is None:
                missing = True
                break
            total += abs(d)
        if missing:
            continue
        scored.append((r["id"], total, r["title"][:60]))
    scored.sort(key=lambda x: -x[1])
    return scored[:n]


# ---- XLSX WRITER ------------------------------------------------------------

BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")   # light blue
STRESS_FILL = PatternFill("solid", start_color="FFF2CC")   # light yellow
OVERHEAD_FILL = PatternFill("solid", start_color="FCE4D6") # light orange
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def write_sidebyside(wb: Workbook, rows: list[dict]):
    ws = wb.active
    ws.title = "Side-by-side"

    headers = [
        "id", "domain", "title", "notes",
        "H-mat", "L-mat", "Δ-mat",
        "H-commerc", "L-profit", "Δ",
        "H-retail", "L-retail", "Δ",
        "H-spec", "L-spec", "Δ",
        "H-final", "L-final", "Δ-final",
        "LLM flag", "LLM TTT",
        "LLM rationale", "Operator rationale",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for r in rows:
        ws.append([
            r["id"], r["domain"], r["title"], r["notes"],
            r["hand_maturation"], r["llm_maturation"], r["delta_maturation"],
            r["hand_profit_mech"], r["llm_profit_mech"], r["delta_profit_mech"],
            r["hand_retail_access"], r["llm_retail_access"], r["delta_retail_access"],
            r["hand_specificity"], r["llm_specificity"], r["delta_specificity"],
            r["hand_final"], r["llm_final"], r["delta_final"],
            r["llm_flag"], r["llm_time_to_thesis"],
            r["llm_rationale"], r["hand_rationale"],
        ])

    # Column widths — keep title + rationales readable
    widths = [5, 9, 42, 16,
              7, 7, 7,
              9, 9, 7,
              8, 8, 7,
              7, 7, 7,
              8, 8, 7,
              10, 9,
              55, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Wrap the long text columns
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (3, 22, 23):  # title, LLM rationale, operator rationale
            ws.cell(row=row_idx, column=col_idx).alignment = WRAP
        # Tint rows by note type
        note = ws.cell(row=row_idx, column=4).value or ""
        if "stress-test" in note:
            for col_idx in range(1, len(headers) + 1):
                if not ws.cell(row=row_idx, column=col_idx).fill.start_color.rgb or \
                   ws.cell(row=row_idx, column=col_idx).fill.start_color.rgb == "00000000":
                    ws.cell(row=row_idx, column=col_idx).fill = STRESS_FILL
        elif "over-operator" in note:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = OVERHEAD_FILL

    # Color-scale the delta columns to highlight divergences (red=negative, green=positive)
    delta_cols = [7, 10, 13, 16, 19]  # column letters: G, J, M, P, S
    for col_idx in delta_cols:
        letter = get_column_letter(col_idx)
        rng = f"{letter}2:{letter}{ws.max_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=-5, start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=5, end_color="63BE7B",
        )
        ws.conditional_formatting.add(rng, rule)

    ws.freeze_panes = "E2"


def write_summary(wb: Workbook, rows: list[dict], stats: dict, top_div: list):
    ws = wb.create_sheet("Divergence summary")

    # Section 1: per-dimension stats
    ws["A1"] = "Per-dimension divergence (excluding papers #2, #3 flagged 'over head')"
    ws["A1"].font = BOLD
    ws.append([])
    ws.append(["Dimension", "n", "mean Δ (LLM − hand)", "mean |Δ|", "min Δ", "max Δ"])
    _style_header(ws, 3, 6)

    def fmt(x):
        return round(x, 2) if x is not None else None

    for _, _, label in DIMS:
        s = stats[label]
        ws.append([
            label, s["n"],
            fmt(s.get("mean_delta")),
            fmt(s.get("mean_abs_delta")),
            fmt(s.get("min_delta")),
            fmt(s.get("max_delta")),
        ])

    # Section 2: top divergent papers
    ws.append([])
    ws.append(["Top papers by total |Δ| across all 4 sub-dims + final"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["id", "total |Δ|", "title"])
    _style_header(ws, ws.max_row, 3)
    for pid, total, title in top_div:
        ws.append([pid, round(total, 2), title])

    # Section 3: flag comparison
    ws.append([])
    ws.append(["Flag behavior"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["LLM flag", "count"])
    _style_header(ws, ws.max_row, 2)
    flag_counts = {}
    for r in rows:
        flag_counts[r["llm_flag"]] = flag_counts.get(r["llm_flag"], 0) + 1
    for f, c in sorted(flag_counts.items()):
        ws.append([f, c])

    # Section 4: stress-test specific
    ws.append([])
    ws.append(["Stress-test papers (charter-designated calibration anchors)"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["id", "title", "H-profit", "L-profit", "Δ", "H-retail", "L-retail", "Δ",
               "H-final", "L-final", "Δ-final", "LLM flag"])
    _style_header(ws, ws.max_row, 12)
    for r in rows:
        if r["id"] not in STRESS_TEST_IDS:
            continue
        ws.append([
            r["id"], r["title"][:50],
            r["hand_profit_mech"], r["llm_profit_mech"], r["delta_profit_mech"],
            r["hand_retail_access"], r["llm_retail_access"], r["delta_retail_access"],
            r["hand_final"], r["llm_final"], r["delta_final"],
            r["llm_flag"],
        ])

    # Column widths
    widths = [26, 10, 12, 12, 10, 10, 10, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_translations(wb: Workbook, rows: list[dict]):
    """The translation field is 'the most important field' per the prompt.
    Give it a dedicated sheet so it's easy to skim."""
    ws = wb.create_sheet("Translations")
    ws.append(["id", "domain", "title", "LLM flag", "LLM public vehicles", "LLM translation"])
    _style_header(ws, 1, 6)

    for r in rows:
        try:
            vehicles = ", ".join(json.loads(r["llm_public_vehicles"])) or "—"
        except json.JSONDecodeError:
            vehicles = r["llm_public_vehicles"]
        ws.append([r["id"], r["domain"], r["title"], r["llm_flag"],
                   vehicles, r["llm_translation"]])

    for i, w in enumerate([5, 9, 42, 10, 18, 80], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Wrap
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).alignment = WRAP
        ws.cell(row=row_idx, column=6).alignment = WRAP
    ws.freeze_panes = "A2"


# ---- MAIN -------------------------------------------------------------------

def main():
    args = parse_args()

    hand = load_csv_by_id(args.hand)
    llm = load_csv_by_id(args.llm)

    rows = compute_divergences(hand, llm)
    stats = aggregate_stats(rows, exclude_ids=OVER_HEAD_IDS)
    top_div = top_divergences(rows, n=6)

    wb = Workbook()
    write_sidebyside(wb, rows)
    write_summary(wb, rows, stats, top_div)
    write_translations(wb, rows)

    wb.save(args.out)

    # Console summary
    print(f"Wrote {args.out}")
    print(f"Papers compared: {len(rows)} ({len(OVER_HEAD_IDS)} excluded from means as 'over-head')")
    print("\nPer-dimension mean |Δ| (LLM − hand), excluding over-head papers:")
    for _, _, label in DIMS:
        s = stats[label]
        if s["mean_abs_delta"] is None:
            continue
        print(f"  {label:<14}  mean Δ={s['mean_delta']:+.2f}  |Δ|={s['mean_abs_delta']:.2f}  "
              f"range=[{s['min_delta']:+.1f}, {s['max_delta']:+.1f}]  (n={s['n']})")
    print(f"\nTop divergent papers:")
    for pid, total, title in top_div:
        print(f"  #{pid:>2} (|Δ|={total:.1f}): {title}")


if __name__ == "__main__":
    main()
